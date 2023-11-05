from httpx import HTTPStatusError
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string
from pydantic import ValidationError

from address import Address
from checklist_sheet import GeneralSheet, IndividualSheet, EntitySheet, reset_cache
from client import Client
from client_dadata import dadata
from errors import convert_errors, mark_error
from meter import Meter
from utils import (
    get_install_date,
    use_ip,
    get_type_client
)

# if dadata.get_balance() <= 300.0:
#     print(f'Необходимо пополнить баланс. На счету {dadata.get_balance()} руб.')

emis_book = load_workbook('./data/checklist/Реестр.xlsx', data_only=True)
ws_emis = emis_book.active

note_column = column_index_from_string('S')  # row[18]

title_row = int(ws_emis.print_title_rows[-int((len(ws_emis.print_title_rows) - 3) / 2):])
note_cell = ws_emis.cell(title_row, note_column, 'Комментарии')
note_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
note_cell.font = Font(bold=True)

quest_book = load_workbook('./data/checklist/template/ОЛ.xlsx', data_only=True)
gener_sheet = quest_book['ТУ']
ind_sheet = quest_book['ФЛ']
ent_sheet = quest_book['ЮЛ']
gener_sheet._current_row = 3  # корр
ind_sheet._current_row = 2  # корр
ent_sheet._current_row = 2  # корр

current_row = title_row + 1  # 1-я строка с данными  -> 7
rows = ws_emis.iter_rows(min_row=current_row)

for row in rows:
    columns = ('D', 'E', 'F', 'G', 'K', 'L', 'R', 'J')  # столбцы с рабочими данными
    num_col = map(lambda i: column_index_from_string(i) - 1, columns)
    cells = map(lambda i: row[i].value, num_col)

    name, account, address, type_client, num_meter, ser_sim, muster_date, act_com = cells

    name, account, address, type_client, num_meter, ser_sim, act_com = \
        (str(i) if i else i for i in (name, account, address, type_client, num_meter, ser_sim, act_com))

    type_client = get_type_client(type_client)

    try:
        install_date = get_install_date(act_com)
    except Exception:
        install_date = None

    note_cell = row[column_index_from_string('S') - 1]
    note_cell.value = str()

    try:
        meter = Meter(num=num_meter, muster_date=muster_date, install_date=install_date, ip=ser_sim)
    except ValidationError as e:
        errors = convert_errors(e.errors())
        mark_error(row, errors)
        meter = None

    # try:
    #     adr = Address(raw_adr=address)
    # except HTTPStatusError as e:
    #     print(e.response.text)
    #     address = None
    #     flat = None
    # except ValidationError as e:
    #     errors = convert_errors(e.errors())
    #     mark_error(row, errors)
    #     address = None
    #     flat = None
    # else:
    #     address, flat = adr.address, adr.flat
    address = 'Магнитогорск, д 1'
    flat = '1'

    try:
        client = Client(type=type_client, name=name, address=address, flat=flat, account=account)
    except ValidationError as e:
        errors = convert_errors(e.errors())
        mark_error(row, errors)
        client = None

    if all((meter, client)):
        # use_ip(ser_sim)
        current_row += 1
        gener_sheet.append(GeneralSheet(meter=meter, client=client).write())

        if client.type == 'ЮЛ':
            try:
                ent_sheet.append(EntitySheet(client=client).write())
            except ValidationError:
                pass
        else:
            try:
                ind_sheet.append(IndividualSheet(client=client).write())
            except ValidationError:
                pass


quest_book.save(f'./data/checklist/ОЛ_{GeneralSheet.number}.xlsx')
emis_book.save('./data/checklist/Реестр.xlsx')
